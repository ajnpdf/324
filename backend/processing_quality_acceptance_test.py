from __future__ import annotations
import tempfile
import zipfile
from pathlib import Path
import fitz
from docx import Document
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from app.conversion_engine import SPECS, validate_output_file
from app.processing_quality import run_conversion

def _font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for candidate in ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', '/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf', 'C:/Windows/Fonts/arial.ttf'):
        try:
            return ImageFont.truetype(candidate, size)
        except OSError:
            continue
    return ImageFont.load_default()

def _digital_pdf(path: Path) -> None:
    doc = fitz.open()
    try:
        page = doc.new_page(width=595, height=842)
        page.insert_text((60, 70), 'AJN PDF Quality Report', fontsize=22)
        page.insert_text((60, 110), 'Meaningful document conversion test.', fontsize=12)
        page.insert_text((60, 135), 'This paragraph should remain readable in Word and Text outputs.', fontsize=11)
        x0, y0 = (60, 200)
        widths = [180, 120, 120]
        row_h = 34
        rows = [['Product', 'Count', 'Amount'], ['AJN PDF', '12', '480']]
        x_positions = [x0]
        for width in widths:
            x_positions.append(x_positions[-1] + width)
        for row_index in range(len(rows) + 1):
            y = y0 + row_index * row_h
            page.draw_line((x0, y), (x_positions[-1], y), width=1)
        for x in x_positions:
            page.draw_line((x, y0), (x, y0 + len(rows) * row_h), width=1)
        for row_index, row in enumerate(rows):
            y = y0 + row_index * row_h + 22
            for column_index, value in enumerate(row):
                page.insert_text((x_positions[column_index] + 8, y), value, fontsize=10)
        page2 = doc.new_page(width=595, height=842)
        page2.insert_text((60, 80), 'Second Page Marker', fontsize=18)
        page2.insert_text((60, 120), 'Page-range image export should select only this page.', fontsize=11)
        doc.save(path)
    finally:
        doc.close()

def _run(tool_id: str, inputs: list[Path], output: Path, workdir: Path, options: dict | None=None) -> Path:
    spec = SPECS[tool_id]
    run_conversion(spec, inputs, output, options or {'language': 'eng', 'dpi': 180, 'quality': 85}, workdir, None)
    validate_output_file(output, spec.output_extension)
    return output

def main() -> None:
    with tempfile.TemporaryDirectory(prefix='ajn-quality-acceptance-') as temporary:
        root = Path(temporary)
        digital = root / 'digital.pdf'
        image = root / 'recognized_text.png'
        _digital_pdf(digital)
        _recognition_image(image)
        text_dir = root / 'text'
        text_dir.mkdir()
        text_output = _run('pdf-to-txt', [digital], text_dir / 'result.txt', text_dir)
        text = text_output.read_text(encoding='utf-8')
        assert 'AJN PDF Quality Report' in text
        assert 'Second Page Marker' in text
        word_dir = root / 'word'
        word_dir.mkdir()
        word_output = _run('pdf-to-word', [digital], word_dir / 'result.docx', word_dir)
        word = Document(word_output)
        word_text = ''.join((paragraph.text for paragraph in word.paragraphs))
        assert 'AJN PDF Quality Report' in word_text
        assert 'Meaningful document conversion test' in word_text
        excel_dir = root / 'excel'
        excel_dir.mkdir()
        excel_output = _run('pdf-to-excel', [digital], excel_dir / 'result.xlsx', excel_dir)
        workbook = load_workbook(excel_output, data_only=True)
        try:
            values = [str(cell.value or '') for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row]
            joined = '|'.join(values)
            assert 'Product' in joined
            assert 'AJN PDF' in joined
            assert '480' in joined
        finally:
            workbook.close()
        image_pdf_dir = root / 'image-pdf'
        image_pdf_dir.mkdir()
        image_pdf = _run('image-to-pdf', [image], image_pdf_dir / 'result.pdf', image_pdf_dir, {'page_size': 'a4', 'margin_mm': 8, 'quality': 88, 'dpi': 160})
        with fitz.open(image_pdf) as result:
            assert result.page_count == 1
            assert result[0].rect.width > 500
        recognition_dir = root / 'recognized_text'
        recognition_dir.mkdir()
        recognition_output = _run('', [image], recognition_dir / 'result.txt', recognition_dir, {'language': 'eng', 'dpi': 240, 'auto_rotate': True, 'deskew': True, 'denoise': True})
        recognized = recognition_output.read_text(encoding='utf-8').upper()
        assert 'AJN' in recognized
        assert '' in recognized
        assert '123' in recognized
        searchable_dir = root / 'searchable'
        searchable_dir.mkdir()
        searchable = _run('', [image], searchable_dir / 'result.pdf', searchable_dir, {'language': 'eng', 'dpi': 220, 'auto_rotate': True, 'deskew': True})
        with fitz.open(searchable) as result:
            searchable_text = ''.join((page.get_text() for page in result)).upper()
            assert 'AJN' in searchable_text
            assert '' in searchable_text
        image_dir = root / 'pdf-image'
        image_dir.mkdir()
        zip_output = _run('pdf-to-png', [digital], image_dir / 'result.zip', image_dir, {'pages': '2', 'dpi': 110, 'quality': 85})
        with zipfile.ZipFile(zip_output) as archive:
            names = archive.namelist()
            assert names == ['page-002.png'], names
        plain_pdf = root / 'plain.pdf'
        plain = fitz.open()
        try:
            page = plain.new_page()
            page.insert_text((72, 72), 'This PDF contains prose but no table.', fontsize=12)
            plain.save(plain_pdf)
        finally:
            plain.close()
        failure_dir = root / 'no-table'
        failure_dir.mkdir()
        try:
            _run('pdf-to-excel', [plain_pdf], failure_dir / 'result.xlsx', failure_dir)
        except ValueError as exc:
            assert 'No structured table' in str(exc)
        else:
            raise AssertionError('PDF to Excel must not generate an unstructured line-dump workbook by default.')
        print('PASS: AJN PDF processing quality acceptance — semantic Word/Excel/Text, searchable PDF, image PDF, page-range export')
if __name__ == '__main__':
    main()
